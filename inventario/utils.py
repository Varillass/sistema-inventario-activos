from io import BytesIO
from datetime import datetime
from django.db.models import Count
from .models import Equipo, Area, Estado, Sede
# import pandas as pd  # Comentado temporalmente para Render
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generar_pdf_equipos(equipos):
    """
    Genera un PDF profesional con la lista de equipos
    """
    # Importar reportlab solo cuando se necesita
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    
    # Crear buffer para el PDF
    buffer = BytesIO()
    
    # Crear documento
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.grey
    )
    
    # Título del documento
    title = Paragraph("INVENTARIO DE ACTIVOS", title_style)
    elements.append(title)
    
    # Subtítulo con fecha
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    subtitle = Paragraph(f"Reporte generado el {fecha}", subtitle_style)
    elements.append(subtitle)
    
    elements.append(Spacer(1, 20))
    
    # Datos de la tabla
    data = [['Nombre', 'Tipo', 'N° Serie', 'Marca/Modelo', 'Precio', 'Área', 'Estado', 'Garantía']]
    
    for equipo in equipos:
        # Formatear precio
        precio_str = f"${equipo.precio:,.2f}" if equipo.precio else "N/A"
        
        # Formatear marca/modelo
        marca_modelo = f"{equipo.marca} {equipo.modelo}".strip()
        if not marca_modelo:
            marca_modelo = "N/A"
        
        # Estado de garantía
        garantia_status = "Vigente" if equipo.garantia_vigente else "Vencida" if equipo.garantia_hasta else "N/A"
        
        data.append([
            equipo.nombre,
            equipo.tipo,
            equipo.numero_serie,
            marca_modelo,
            precio_str,
            equipo.area.nombre,
            equipo.estado.nombre,
            garantia_status
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[1.5*inch, 0.8*inch, 1*inch, 1.2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch])
    
    # Estilo de la tabla
    style = TableStyle([
        # Encabezados
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        
        # Datos
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        
        # Alineación específica para columnas
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),  # Precio
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # N° Serie
        ('ALIGN', (6, 1), (6, -1), 'CENTER'),  # Estado
        ('ALIGN', (7, 1), (7, -1), 'CENTER'),  # Garantía
    ])
    
    table.setStyle(style)
    elements.append(table)
    
    # Estadísticas
    elements.append(Spacer(1, 30))
    
    # Contar equipos por estado
    estados_count = {}
    for equipo in equipos:
        estado = equipo.estado.nombre
        estados_count[estado] = estados_count.get(estado, 0) + 1
    
    # Crear tabla de estadísticas
    stats_data = [['Estadísticas del Inventario']]
    stats_data.append(['Estado', 'Cantidad'])
    
    for estado, cantidad in estados_count.items():
        stats_data.append([estado, str(cantidad)])
    
    stats_table = Table(stats_data, colWidths=[2*inch, 1*inch])
    stats_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgreen]),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ])
    
    stats_table.setStyle(stats_style)
    elements.append(stats_table)
    
    # Construir PDF
    doc.build(elements)
    
    # Obtener PDF
    pdf = buffer.getvalue()
    buffer.close()
    
    return pdf

def generar_excel_equipos(equipos):
    """
    Genera un archivo Excel profesional con la lista de equipos
    """
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario de Activos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    data_font = Font(size=10)
    data_alignment = Alignment(horizontal="left", vertical="center")
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = "INVENTARIO DE ACTIVOS"
    ws['A1'].font = Font(bold=True, size=16, color="366092")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Fecha
    ws.merge_cells('A2:H2')
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws['A2'] = f"Reporte generado el {fecha}"
    ws['A2'].font = Font(size=10, color="666666")
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Encabezados
    headers = [
        'Nombre', 'Tipo', 'N° Serie', 'Marca', 'Modelo', 'Precio', 
        'Proveedor', 'Fecha Compra', 'Garantía Hasta', 'Área', 'Estado', 'Observación'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Datos
    for row, equipo in enumerate(equipos, 5):
        # Formatear fecha de compra
        fecha_compra = equipo.fecha_compra.strftime("%d/%m/%Y") if equipo.fecha_compra else ""
        
        # Formatear garantía
        garantia_hasta = equipo.garantia_hasta.strftime("%d/%m/%Y") if equipo.garantia_hasta else ""
        
        # Formatear precio
        precio = f"${equipo.precio:,.2f}" if equipo.precio else ""
        
        # Datos de la fila
        row_data = [
            equipo.nombre,
            equipo.tipo,
            equipo.numero_serie,
            equipo.marca or "",
            equipo.modelo or "",
            precio,
            equipo.proveedor or "",
            fecha_compra,
            garantia_hasta,
            equipo.area.nombre,
            equipo.estado.nombre,
            equipo.observacion or ""
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # Ajustar ancho de columnas
    column_widths = [25, 15, 15, 15, 15, 12, 20, 12, 12, 15, 15, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Crear hoja de estadísticas
    ws_stats = wb.create_sheet("Estadísticas")
    
    # Estadísticas por estado
    estados_count = {}
    for equipo in equipos:
        estado = equipo.estado.nombre
        estados_count[estado] = estados_count.get(estado, 0) + 1
    
    ws_stats['A1'] = "Estadísticas del Inventario"
    ws_stats['A1'].font = Font(bold=True, size=14, color="366092")
    
    ws_stats['A3'] = "Estado"
    ws_stats['B3'] = "Cantidad"
    ws_stats['A3'].font = header_font
    ws_stats['B3'].font = header_font
    ws_stats['A3'].fill = header_fill
    ws_stats['B3'].fill = header_fill
    
    row = 4
    for estado, cantidad in estados_count.items():
        ws_stats[f'A{row}'] = estado
        ws_stats[f'B{row}'] = cantidad
        row += 1
    
    # Ajustar ancho de columnas en estadísticas
    ws_stats.column_dimensions['A'].width = 20
    ws_stats.column_dimensions['B'].width = 10
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()

def generar_plantilla_excel():
    """
    Genera una plantilla Excel para importar equipos
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Importación"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Encabezados
    headers = [
        'Nombre*', 'Tipo*', 'N° Serie', 'Marca', 'Modelo', 'Precio', 
        'Proveedor', 'Fecha Compra (DD/MM/YYYY)', 'Garantía Hasta (DD/MM/YYYY)', 
        'Sede*', 'Área*', 'Estado*', 'Observación'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Ejemplos de datos
    ejemplo_data = [
        'Computadora Dell', 'Computadora', 'Com-00001', 'Dell', 'OptiPlex 7090', '1200.00',
        'Dell Technologies', '15/01/2023', '15/01/2026', 'Sede Principal', 'Administración', 'Operativo', 'Equipo principal'
    ]
    
    for col, value in enumerate(ejemplo_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    # Ajustar ancho de columnas
    column_widths = [25, 15, 15, 15, 15, 12, 20, 25, 25, 15, 15, 15, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()

def validar_archivo_excel(archivo):
    """
    Valida que el archivo Excel tenga el formato correcto
    """
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        # Verificar que tenga al menos una fila de datos
        if ws.max_row < 2:
            return False, "El archivo no contiene datos"
        
        # Obtener headers y limpiarlos
        headers = []
        for cell in ws[1]:
            header_value = str(cell.value).strip() if cell.value else ''
            # Limpiar asteriscos y caracteres especiales
            header_value = header_value.replace('*', '').replace('(', '').replace(')', '').strip()
            headers.append(header_value)
        
        # Verificar columnas requeridas con mapeo flexible
        columnas_requeridas = {
            'nombre': ['nombre', 'name', 'equipo', 'equipment'],
            'tipo': ['tipo', 'type', 'categoria', 'category'],
            'sede': ['sede', 'location', 'ubicacion', 'site'],
            'area': ['area', 'área', 'departamento', 'department'],
            'estado': ['estado', 'status', 'condition'],
        }
        
        headers_lower = [h.lower() for h in headers]
        
        for campo, posibles_nombres in columnas_requeridas.items():
            encontrada = False
            for header_lower in headers_lower:
                if any(nombre in header_lower for nombre in posibles_nombres):
                    encontrada = True
                    break
            
            if not encontrada:
                return False, f"Falta la columna requerida: {campo}. Busque: {', '.join(posibles_nombres)}"
        
        return True, "Archivo válido"
        
    except Exception as e:
        return False, f"Error al leer el archivo: {str(e)}"

def procesar_importacion_excel(archivo):
    """
    Procesa la importación masiva desde Excel usando openpyxl
    """
    try:
        # Cargar archivo Excel
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        # Obtener encabezados de la primera fila y limpiarlos
        headers = []
        for cell in ws[1]:
            header_value = str(cell.value).strip() if cell.value else ''
            # Limpiar asteriscos y caracteres especiales
            header_value = header_value.replace('*', '').replace('(', '').replace(')', '').strip()
            headers.append(header_value)
        
        # Mapeo de columnas esperadas
        columnas_requeridas = {
            'nombre': ['nombre', 'name', 'equipo', 'equipment'],
            'tipo': ['tipo', 'type', 'categoria', 'category'],
            'sede': ['sede', 'location', 'ubicacion', 'site'],
            'area': ['area', 'área', 'departamento', 'department'],
            'estado': ['estado', 'status', 'condition'],
        }
        
        columnas_opcionales = {
            'marca': ['marca', 'brand', 'fabricante', 'manufacturer'],
            'modelo': ['modelo', 'model'],
            'precio': ['precio', 'price', 'cost', 'costo', 'valor', 'value'],
            'proveedor': ['proveedor', 'supplier', 'vendor'],
            'observacion': ['observacion', 'observación', 'descripcion', 'descripción', 'description', 'notes', 'notas'],
            'fecha_compra': ['fecha_compra', 'fecha compra', 'purchase_date', 'date_purchased', 'compra'],
            'garantia_hasta': ['garantia_hasta', 'garantía hasta', 'warranty_until', 'garantia', 'garantía'],
            'numero_serie': ['numero_serie', 'número de serie', 'serial', 'serie', 'n° serie', 'n serie']
        }
        
        # Mapear columnas del archivo a nuestros campos
        mapeo_columnas = {}
        
        # Buscar columnas requeridas
        for campo, posibles_nombres in columnas_requeridas.items():
            encontrada = False
            for i, header in enumerate(headers):
                if any(nombre.lower() in header.lower() for nombre in posibles_nombres):
                    mapeo_columnas[campo] = i
                    encontrada = True
                    break
            
            if not encontrada:
                return 0, [f"Columna requerida '{campo}' no encontrada. Busque: {', '.join(posibles_nombres)}"]
        
        # Buscar columnas opcionales
        for campo, posibles_nombres in columnas_opcionales.items():
            for i, header in enumerate(headers):
                if any(nombre.lower() in header.lower() for nombre in posibles_nombres):
                    mapeo_columnas[campo] = i
                    break
        
        # Procesar datos
        equipos_creados = 0
        errores = []
        
        for row_num in range(2, ws.max_row + 1):  # Empezar desde fila 2 (después de headers)
            try:
                # Extraer datos requeridos
                nombre = str(ws.cell(row=row_num, column=mapeo_columnas['nombre'] + 1).value or '').strip()
                tipo = str(ws.cell(row=row_num, column=mapeo_columnas['tipo'] + 1).value or '').strip()
                sede_nombre = str(ws.cell(row=row_num, column=mapeo_columnas['sede'] + 1).value or '').strip()
                area_nombre = str(ws.cell(row=row_num, column=mapeo_columnas['area'] + 1).value or '').strip()
                estado_nombre = str(ws.cell(row=row_num, column=mapeo_columnas['estado'] + 1).value or '').strip()
                
                # Validar datos requeridos
                if not nombre or nombre == 'None':
                    errores.append(f"Fila {row_num}: Nombre es requerido")
                    continue
                
                if not tipo or tipo == 'None':
                    errores.append(f"Fila {row_num}: Tipo es requerido")
                    continue
                
                if not sede_nombre or sede_nombre == 'None':
                    errores.append(f"Fila {row_num}: Sede es requerida")
                    continue
                
                # Buscar o crear sede
                try:
                    sede = Sede.objects.get(nombre__iexact=sede_nombre)
                except Sede.DoesNotExist:
                    sede = Sede.objects.create(nombre=sede_nombre)
                
                # Buscar o crear área
                try:
                    area = Area.objects.get(nombre__iexact=area_nombre)
                except Area.DoesNotExist:
                    area = Area.objects.create(nombre=area_nombre)
                
                # Buscar o crear estado
                try:
                    estado = Estado.objects.get(nombre__iexact=estado_nombre)
                except Estado.DoesNotExist:
                    estado = Estado.objects.create(nombre=estado_nombre)
                
                # Datos opcionales
                marca = str(ws.cell(row=row_num, column=mapeo_columnas.get('marca', 0) + 1).value or '').strip()
                modelo = str(ws.cell(row=row_num, column=mapeo_columnas.get('modelo', 0) + 1).value or '').strip()
                proveedor = str(ws.cell(row=row_num, column=mapeo_columnas.get('proveedor', 0) + 1).value or '').strip()
                observacion = str(ws.cell(row=row_num, column=mapeo_columnas.get('observacion', 0) + 1).value or '').strip()
                
                # Procesar precio
                precio_cell = ws.cell(row=row_num, column=mapeo_columnas.get('precio', 0) + 1)
                precio = None
                if precio_cell.value:
                    try:
                        precio = float(str(precio_cell.value).replace(',', '').replace('$', ''))
                    except (ValueError, TypeError):
                        errores.append(f"Fila {row_num}: Precio inválido")
                        continue
                
                # Procesar fechas
                fecha_compra = None
                garantia_hasta = None
                
                fecha_compra_cell = ws.cell(row=row_num, column=mapeo_columnas.get('fecha_compra', 0) + 1)
                if fecha_compra_cell.value:
                    try:
                        if isinstance(fecha_compra_cell.value, datetime):
                            fecha_compra = fecha_compra_cell.value.date()
                        else:
                            fecha_compra = datetime.strptime(str(fecha_compra_cell.value), '%d/%m/%Y').date()
                    except (ValueError, TypeError):
                        errores.append(f"Fila {row_num}: Fecha de compra inválida")
                        continue
                
                garantia_cell = ws.cell(row=row_num, column=mapeo_columnas.get('garantia_hasta', 0) + 1)
                if garantia_cell.value:
                    try:
                        if isinstance(garantia_cell.value, datetime):
                            garantia_hasta = garantia_cell.value.date()
                        else:
                            garantia_hasta = datetime.strptime(str(garantia_cell.value), '%d/%m/%Y').date()
                    except (ValueError, TypeError):
                        errores.append(f"Fila {row_num}: Fecha de garantía inválida")
                        continue
                
                # Generar número de serie si no existe
                numero_serie = ''
                if 'numero_serie' in mapeo_columnas:
                    numero_serie = str(ws.cell(row=row_num, column=mapeo_columnas['numero_serie'] + 1).value or '').strip()
                
                if not numero_serie:
                    # Generar número de serie automático
                    prefijo = tipo[:3].upper() if len(tipo) >= 3 else tipo.upper()
                    ultimo_equipo = Equipo.objects.filter(numero_serie__startswith=f"{prefijo}-").order_by('-numero_serie').first()
                    
                    if ultimo_equipo:
                        try:
                            ultimo_numero = int(ultimo_equipo.numero_serie.split('-')[-1])
                            numero_serie = f"{prefijo}-{ultimo_numero + 1:05d}"
                        except (ValueError, IndexError):
                            numero_serie = f"{prefijo}-00001"
                    else:
                        numero_serie = f"{prefijo}-00001"
                
                # Crear equipo
                equipo = Equipo.objects.create(
                    nombre=nombre,
                    tipo=tipo,
                    numero_serie=numero_serie,
                    marca=marca,
                    modelo=modelo,
                    precio=precio,
                    proveedor=proveedor,
                    fecha_compra=fecha_compra,
                    garantia_hasta=garantia_hasta,
                    observacion=observacion,
                    sede=sede,
                    area=area,
                    estado=estado
                )
                
                equipos_creados += 1
                
            except Exception as e:
                # Error más específico
                error_msg = str(e)
                if "DoesNotExist" in error_msg:
                    errores.append(f"Fila {row_num}: Error de referencia - verifica que los datos existan")
                elif "ValidationError" in error_msg:
                    errores.append(f"Fila {row_num}: Error de validación - verifica el formato de los datos")
                else:
                    errores.append(f"Fila {row_num}: Error - {error_msg}")
                continue
        
        return equipos_creados, errores
        
    except Exception as e:
        return 0, [f"Error al procesar archivo: {str(e)}"]

def crear_equipos_masivo(equipos_data):
    """
    Crea los equipos en la base de datos de forma masiva
    """
    equipos_creados = []
    errores = []
    
    for equipo_data in equipos_data:
        try:
            fila = equipo_data.pop('fila')
            
            # Verificar si ya existe un equipo con el mismo nombre en la misma área
            if Equipo.objects.filter(
                nombre__iexact=equipo_data['nombre'], 
                area=equipo_data['area']
            ).exists():
                errores.append(f"Fila {fila}: Ya existe un equipo '{equipo_data['nombre']}' en el área '{equipo_data['area'].nombre}'")
                continue
            
            # Crear el equipo
            equipo = Equipo.objects.create(**equipo_data)
            equipos_creados.append(equipo)
            
        except Exception as e:
            errores.append(f"Fila {fila}: Error al crear equipo - {str(e)}")
    
    return equipos_creados, errores 